"""
Generates the Case Management demo:

  case_management_example_form.xlsx  — XLSForm to deploy in KoboToolbox
  case_management_example_cases.csv  — case table CSV to upload as a Case Table

Flow: you type a **Case ID**. The form looks it up in the linked case table
and tells you whether it is new or already on file. Either way the Name and
Status fields are **pre-filled with what is on file and stay editable**, so an
existing case can be corrected and a new one filled in from blank. Submitting
writes the values back onto the case record, creating it if the id was unknown.

The pre-filling trick: `trigger`
--------------------------------
A `default` is evaluated once when the form loads, which is too early — the
case id has not been typed yet. Instead these questions carry a `calculation`
*and* a `trigger` pointing at `${case_id}`. pyxform turns that into

    <setvalue event="xforms-value-changed" ref="/data/name" value="pulldata(…)"/>

attached to the case id question, so every time the case id changes the field
is re-populated from the case table — while remaining an ordinary editable
question. That is what makes "load the existing name, but let me change it"
possible.

Because Name is always present in the submission, the write-back always writes
it. That is intended here (it is pre-filled with the stored value, so an
unchanged submission is a no-op) — but note that deliberately clearing the
field will blank the stored name. Where a field must never be overwritten by a
blank, gate it with `relevant` instead: a non-relevant question is absent from
the submission and is skipped by the write-back.

The last question is a `select_one_from_file cases.csv` browser. It is not
mapped to any column and does not affect the case record — it is there to show
that reading the case table as a choice list works too, with the choices coming
live from the table.

Link configuration this form expects (project Settings → Case Management):

    Attached file name      cases.csv
    Case id question name   case_id
    Write-back mappings     name=name
                            status=status
                            visit_date=last_visit
    Write submissions back to cases        checked
    Create unknown cases automatically     checked
"""
import csv
import os

import openpyxl

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

wb = openpyxl.Workbook()

# --- survey sheet ---
survey = wb.active
survey.title = 'survey'
survey.append([
    'type', 'name', 'label', 'hint', 'calculation', 'trigger', 'relevant',
    'required', 'parameters', 'appearance',
])

survey.append([
    'text', 'case_id', 'Case ID',
    'Type an id that exists (CASE001) or a brand new one (CASE900).',
    '', '', '', 'yes', '', '',
])

# Everything on file for that id. Empty strings when the id is unknown.
survey.append([
    'calculate', 'known_name', '', '',
    "pulldata('cases', 'name', 'case_id', ${case_id})", '', '', '', '', '',
])
survey.append([
    'calculate', 'known_status', '', '',
    "pulldata('cases', 'status', 'case_id', ${case_id})", '', '', '', '', '',
])
survey.append([
    'calculate', 'known_last_visit', '', '',
    "pulldata('cases', 'last_visit', 'case_id', ${case_id})", '', '', '', '',
    '',
])
survey.append([
    'calculate', 'case_found', '', '',
    "if(string-length(${known_name}) > 0 or string-length(${known_status}) > 0 "
    "or string-length(${known_last_visit}) > 0, 'yes', 'no')",
    '', '', '', '', '',
])

# Tell the enumerator which of the two situations they are in.
survey.append([
    'note', 'existing_case',
    '**Existing case — ${case_id}**\n'
    '- Name on file: ${known_name}\n'
    '- Status on file: ${known_status}\n'
    '- Last visit on file: ${known_last_visit}\n\n'
    'The fields below are pre-filled; change what you need.',
    '', '', '', "${case_found} = 'yes'", '', '', '',
])
survey.append([
    'note', 'new_case',
    '**NEW — ${case_id} is not in the case table**\n\n'
    'It will be created when you submit. Fill in the name below.',
    '', '', '', "${case_found} = 'no'", '', '', '',
])

# Pre-filled from the case table on every change of ${case_id}, still editable.
survey.append([
    'text', 'name', 'Name',
    'Pre-filled from the case table for a known case; type it in for a new one.',
    "pulldata('cases', 'name', 'case_id', ${case_id})", '${case_id}', '',
    'yes', '', '',
])
survey.append([
    'select_one status_list', 'status', 'Status',
    'Pre-filled with the status on file.',
    "pulldata('cases', 'status', 'case_id', ${case_id})", '${case_id}', '',
    'yes', '', 'minimal',
])
survey.append([
    'date', 'visit_date', 'Visit date', '', '', '', '', 'yes', '', '',
])
survey.append([
    'text', 'notes', 'Notes (not written back)', '', '', '', '', '', '', '',
])

# --- Demonstration of select_one_from_file -------------------------------
# Not mapped to any column: picking something here does not touch the case
# record. It only shows that the same live file also works as a choice list.
survey.append([
    'note', 'browser_intro',
    '---\n**Bonus: the same case table as a dropdown**\n\n'
    'The list below is built live from `cases.csv` — the same file '
    '`pulldata()` reads above. It is here only to demonstrate '
    '`select_one_from_file`; it is not written back.',
    '', '', '', '', '', '', '',
])
survey.append([
    'select_one_from_file cases.csv', 'case_browser',
    'Browse existing cases (demo only)', '', '', '', '', '',
    'value=case_id,label=name', 'minimal',
])

# --- choices sheet ---
choices = wb.create_sheet('choices')
choices.append(['list_name', 'name', 'label'])
choices.append(['status_list', 'open', 'Open'])
choices.append(['status_list', 'in_progress', 'In progress'])
choices.append(['status_list', 'closed', 'Closed'])

# --- settings sheet ---
settings = wb.create_sheet('settings')
settings.append(['form_title', 'form_id'])
settings.append(['Case Management Example', 'case_management_example'])

wb.save(f'{OUT_DIR}/case_management_example_form.xlsx')

# --- case table CSV ---
with open(f'{OUT_DIR}/case_management_example_cases.csv', 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['case_id', 'name', 'status', 'last_visit'])
    writer.writerow(['CASE001', 'Amina Diallo', 'open', '2026-08-01'])
    writer.writerow(['CASE002', 'Jean Kouassi', 'in_progress', '2026-08-10'])
    writer.writerow(['CASE003', 'Fatou Sow', 'closed', '2026-07-15'])
    # Any other id (e.g. CASE900) exercises "create unknown cases
    # automatically".

print('done')
